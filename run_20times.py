#!/usr/bin/env python3
"""
Run exact (1 time) + METHOD (20 times) on test data, export to Excel.

Environment variables:
  DATASET   Dataset name (default: tiny)
            Options: tiny, dataset_a, dataset_b, dataset_c, dataset_d
  METHOD    Method to run 20 times (default: v2)
            Options: v2, mcmc
  RUNS      Number of runs (default: 20)
  SAMPLES   MCMC samples per run (default: 10000, only for mcmc)

Usage:
  DATASET=dataset_c METHOD=v2   python3 run_20times.py
  DATASET=dataset_c METHOD=mcmc python3 run_20times.py
  DATASET=dataset_b METHOD=mcmc SAMPLES=20000 python3 run_20times.py
"""

import subprocess
import os

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))

DATASET = os.environ.get("DATASET", "tiny")
METHOD = os.environ.get("METHOD", "v2")
NUM_RUNS = int(os.environ.get("RUNS", "20"))
MCMC_SAMPLES = os.environ.get("SAMPLES", "10000")
NUM_TYPES = 18

RESULTS_DIR = os.path.join(PROJECT_DIR, "results", f"runs_{DATASET}_{METHOD}")

NVERTS = os.path.join(PROJECT_DIR, "testdata", DATASET, "nverts.txt")
SIMPLICES = os.path.join(PROJECT_DIR, "testdata", DATASET, "simplices.txt")
if not os.path.exists(NVERTS):
    NVERTS = os.path.join(PROJECT_DIR, "testdata", "nverts.txt")
    SIMPLICES = os.path.join(PROJECT_DIR, "testdata", "simplices.txt")

METHOD_NAMES = {"v2": "UNIFIED_V2", "mcmc": "MCMC_ORIGINAL"}
METHOD_LABEL = METHOD_NAMES.get(METHOD, METHOD.upper())

os.makedirs(RESULTS_DIR, exist_ok=True)
subprocess.run(["make", "all"], cwd=PROJECT_DIR, capture_output=True)
subprocess.run(["make", "testdata"], cwd=PROJECT_DIR, capture_output=True)

def parse_res(filepath):
    data = {"method": "", "time_ms": 0, "total": 0, "samples": 0,
            "counts": [0]*NUM_TYPES, "freq": [0.0]*NUM_TYPES,
            "ci_lo": [0.0]*NUM_TYPES, "ci_hi": [0.0]*NUM_TYPES}
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if line.startswith("METHOD:"): data["method"] = line.split(":", 1)[1].strip()
            elif line.startswith("TIME_MS:"): data["time_ms"] = float(line.split(":", 1)[1].strip())
            elif line.startswith("TOTAL:"): data["total"] = int(line.split(":", 1)[1].strip())
            elif line.startswith("SAMPLES:"): data["samples"] = int(line.split(":", 1)[1].strip())
            elif line[0:1].isdigit():
                parts = line.split(",")
                if len(parts) >= 3:
                    t = int(parts[0]) - 1
                    if 0 <= t < NUM_TYPES:
                        data["counts"][t] = int(parts[1])
                        data["freq"][t] = float(parts[2])
                        if len(parts) >= 5:
                            data["ci_lo"][t] = float(parts[3])
                            data["ci_hi"][t] = float(parts[4])
    return data

# === RUN EXACT ===
print(f"Dataset: {DATASET}")
print(f"Method:  {METHOD_LABEL} x {NUM_RUNS} runs")
if METHOD == "mcmc": print(f"Samples: {MCMC_SAMPLES} per run")
print()

print("Running exact enumeration...")
exact_file = os.path.join(RESULTS_DIR, "exact.res")
subprocess.run([os.path.join(PROJECT_DIR, "exact_sfd"), NVERTS, SIMPLICES, exact_file],
               cwd=PROJECT_DIR, capture_output=True)
exact = parse_res(exact_file)
print(f"  Exact: {exact['total']} simplets, {exact['time_ms']:.3f} ms")

# === RUN METHOD x N ===
runs = []
for i in range(1, NUM_RUNS + 1):
    print(f"Running {METHOD_LABEL} #{i}/{NUM_RUNS}...", end=" ", flush=True)
    out_file = os.path.join(RESULTS_DIR, f"{METHOD}_run{i}.res")
    if METHOD == "v2":
        subprocess.run([os.path.join(PROJECT_DIR, "unified_sfd_v2"),
                        NVERTS, SIMPLICES, out_file, "0"],
                       cwd=PROJECT_DIR, capture_output=True)
    elif METHOD == "mcmc":
        subprocess.run([os.path.join(PROJECT_DIR, "mcmc_sfd"),
                        NVERTS, SIMPLICES, out_file, "0", MCMC_SAMPLES],
                       cwd=PROJECT_DIR, capture_output=True)
    data = parse_res(out_file)
    runs.append(data)
    print(f"samples={data['samples']}, time={data['time_ms']:.0f}ms")
print(f"\nAll {NUM_RUNS} runs complete.")

# === BUILD EXCEL ===
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
EXACT_FILL = PatternFill("solid", fgColor="D6E4F0")
RUN_FILL = PatternFill("solid", fgColor="E2EFDA") if METHOD == "v2" else PatternFill("solid", fgColor="FDE9D9")
STAT_FILL = PatternFill("solid", fgColor="FFF2CC")
ERR_FILL = PatternFill("solid", fgColor="FCE4EC")
BOLD = Font(name="Arial", bold=True, size=11)
NORMAL = Font(name="Arial", size=11)
SMALL = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

def sc(cell, font=NORMAL, fill=None, align=CENTER, border=THIN, fmt=None):
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = align
    cell.border = border
    if fmt: cell.number_format = fmt

def aw(ws):
    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml + 3, 10), 20)

active_types = [t for t in range(NUM_TYPES)
                if exact["freq"][t] > 0.0001 or any(r["freq"][t] > 0.0001 for r in runs)]
has_ci = (METHOD == "v2")

# --- SHEET 1: Frequencies ---
ws1 = wb.active
ws1.title = f"{METHOD_LABEL} Freq"
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2+len(active_types))
sc(ws1.cell(row=1, column=1, value=f"SFD: Exact vs {METHOD_LABEL} ({NUM_RUNS} runs, {DATASET})"),
   Font(name="Arial", bold=True, size=14, color="2F5496"))

row = 3
sc(ws1.cell(row=row, column=1, value="Run"), HEADER_FONT, HEADER_FILL)
sc(ws1.cell(row=row, column=2, value="Samples"), HEADER_FONT, HEADER_FILL)
for ci, t in enumerate(active_types):
    sc(ws1.cell(row=row, column=3+ci, value=f"Type {t+1}"), HEADER_FONT, HEADER_FILL)
sc(ws1.cell(row=row, column=3+len(active_types), value="Time(ms)"), HEADER_FONT, HEADER_FILL)

row = 4
sc(ws1.cell(row=row, column=1, value="EXACT"), BOLD, EXACT_FILL)
sc(ws1.cell(row=row, column=2, value=exact["total"]), NORMAL, EXACT_FILL)
for ci, t in enumerate(active_types):
    sc(ws1.cell(row=row, column=3+ci, value=exact["freq"][t]), NORMAL, EXACT_FILL, fmt="0.000000")
sc(ws1.cell(row=row, column=3+len(active_types), value=exact["time_ms"]), NORMAL, EXACT_FILL, fmt="0.000")

for i, run in enumerate(runs):
    row = 5 + i
    sc(ws1.cell(row=row, column=1, value=f"#{i+1}"), NORMAL, RUN_FILL)
    sc(ws1.cell(row=row, column=2, value=run["samples"]), NORMAL, RUN_FILL)
    for ci, t in enumerate(active_types):
        sc(ws1.cell(row=row, column=3+ci, value=run["freq"][t]), NORMAL, fmt="0.000000")
    sc(ws1.cell(row=row, column=3+len(active_types), value=run["time_ms"]), NORMAL, fmt="0.000")

sr = 5 + NUM_RUNS + 1
fd, ld = 5, 5 + NUM_RUNS - 1
for si, label in enumerate(["MEAN", "STD", "MIN", "MAX", "EXACT", "MEAN Error"]):
    r = sr + si
    sc(ws1.cell(row=r, column=1, value=label), BOLD, STAT_FILL)
    sc(ws1.cell(row=r, column=2, value=""), NORMAL, STAT_FILL)
    for ci, t in enumerate(active_types):
        cl = get_column_letter(3 + ci)
        c = ws1.cell(row=r, column=3+ci)
        sc(c, NORMAL, STAT_FILL, fmt="0.000000")
        if label == "MEAN": c.value = f"=AVERAGE({cl}{fd}:{cl}{ld})"
        elif label == "STD": c.value = f"=STDEV({cl}{fd}:{cl}{ld})"
        elif label == "MIN": c.value = f"=MIN({cl}{fd}:{cl}{ld})"
        elif label == "MAX": c.value = f"=MAX({cl}{fd}:{cl}{ld})"
        elif label == "EXACT": c.value = exact["freq"][t]
        elif label == "MEAN Error":
            c.value = f"=ABS({cl}{sr}-{cl}{sr+4})"
            sc(c, NORMAL, ERR_FILL, fmt="0.000000")
    tcl = get_column_letter(3+len(active_types))
    tc = ws1.cell(row=r, column=3+len(active_types))
    sc(tc, NORMAL, STAT_FILL, fmt="0.000")
    if label == "MEAN": tc.value = f"=AVERAGE({tcl}{fd}:{tcl}{ld})"
    elif label == "STD": tc.value = f"=STDEV({tcl}{fd}:{tcl}{ld})"
    elif label == "MIN": tc.value = f"=MIN({tcl}{fd}:{tcl}{ld})"
    elif label == "MAX": tc.value = f"=MAX({tcl}{fd}:{tcl}{ld})"
    elif label == "EXACT": tc.value = exact["time_ms"]
    elif label == "MEAN Error": tc.value = ""
aw(ws1)

# --- SHEET 2: Errors ---
ws2 = wb.create_sheet(f"{METHOD_LABEL} Errors")
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2+len(active_types))
sc(ws2.cell(row=1, column=1, value=f"|{METHOD_LABEL} - Exact| per run"),
   Font(name="Arial", bold=True, size=14, color="2F5496"))

row = 3
sc(ws2.cell(row=row, column=1, value="Run"), HEADER_FONT, HEADER_FILL)
for ci, t in enumerate(active_types):
    sc(ws2.cell(row=row, column=2+ci, value=f"Type {t+1}"), HEADER_FONT, HEADER_FILL)
sc(ws2.cell(row=row, column=2+len(active_types), value="RMSE"), HEADER_FONT, HEADER_FILL)
sc(ws2.cell(row=row, column=3+len(active_types), value="MAE"), HEADER_FONT, HEADER_FILL)

for i, run in enumerate(runs):
    row = 4 + i
    sc(ws2.cell(row=row, column=1, value=f"#{i+1}"), NORMAL, RUN_FILL)
    errors = []
    for ci, t in enumerate(active_types):
        err = abs(run["freq"][t] - exact["freq"][t])
        errors.append(err)
        sc(ws2.cell(row=row, column=2+ci, value=err), NORMAL, ERR_FILL if err > 0.01 else None, fmt="0.000000")
    rmse = (sum(e**2 for e in errors) / len(errors)) ** 0.5
    sc(ws2.cell(row=row, column=2+len(active_types), value=rmse), NORMAL, fmt="0.000000")
    sc(ws2.cell(row=row, column=3+len(active_types), value=sum(errors)/len(errors)), NORMAL, fmt="0.000000")

sr2 = 4 + NUM_RUNS + 1
for si, label in enumerate(["MEAN", "STD", "MAX"]):
    r = sr2 + si
    sc(ws2.cell(row=r, column=1, value=label), BOLD, STAT_FILL)
    for ci in range(len(active_types) + 2):
        cl = get_column_letter(2 + ci)
        c = ws2.cell(row=r, column=2+ci)
        sc(c, NORMAL, STAT_FILL, fmt="0.000000")
        if label == "MEAN": c.value = f"=AVERAGE({cl}4:{cl}{4+NUM_RUNS-1})"
        elif label == "STD": c.value = f"=STDEV({cl}4:{cl}{4+NUM_RUNS-1})"
        elif label == "MAX": c.value = f"=MAX({cl}4:{cl}{4+NUM_RUNS-1})"
aw(ws2)

# --- SHEET 3: CI Coverage (V2 only) ---
total_checks, total_covered = 0, 0
if has_ci:
    ws3 = wb.create_sheet("CI Coverage")
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    sc(ws3.cell(row=1, column=1, value="95% CI Coverage"),
       Font(name="Arial", bold=True, size=14, color="2F5496"))
    row = 3
    for h, val in enumerate(["Run", "Type", "Exact", "CI Low", "CI High", "Covered?"]):
        sc(ws3.cell(row=row, column=1+h, value=val), HEADER_FONT, HEADER_FILL)
    row = 4
    for i, run in enumerate(runs):
        for t in active_types:
            ex = exact["freq"][t]
            lo, hi = run["ci_lo"][t], run["ci_hi"][t]
            covered = lo <= ex <= hi
            total_checks += 1
            if covered: total_covered += 1
            sc(ws3.cell(row=row, column=1, value=f"#{i+1}"), SMALL)
            sc(ws3.cell(row=row, column=2, value=f"Type {t+1}"), SMALL)
            sc(ws3.cell(row=row, column=3, value=ex), SMALL, fmt="0.000000")
            sc(ws3.cell(row=row, column=4, value=lo), SMALL, fmt="0.000000")
            sc(ws3.cell(row=row, column=5, value=hi), SMALL, fmt="0.000000")
            fc = PatternFill("solid", fgColor="C6EFCE") if covered else PatternFill("solid", fgColor="FFC7CE")
            fn = Font(name="Arial", bold=True, size=10, color="006100" if covered else "9C0006")
            sc(ws3.cell(row=row, column=6, value="YES" if covered else "NO"), fn, fc)
            row += 1
    row += 1
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    pct = 100.0 * total_covered / total_checks if total_checks > 0 else 0
    sc(ws3.cell(row=row, column=1, value=f"Coverage: {total_covered}/{total_checks} = {pct:.1f}%"),
       Font(name="Arial", bold=True, size=13, color="2F5496"))
    aw(ws3)

# --- SHEET 4: Summary ---
ws4 = wb.create_sheet("Summary")
ws4.sheet_properties.tabColor = "2F5496"
wb.move_sheet("Summary", offset=-(len(wb.sheetnames)-1))

ws4.merge_cells("A1:F1")
sc(ws4.cell(row=1, column=1, value=f"Exact vs {METHOD_LABEL} ({DATASET})"),
   Font(name="Arial", bold=True, size=16, color="2F5496"))

row = 3
info = [("Dataset", DATASET), ("Method", METHOD_LABEL), ("Runs", str(NUM_RUNS)),
        ("Exact Simplets", str(exact["total"]))]
if METHOD == "mcmc":
    info += [("Samples/run", MCMC_SAMPLES), ("Steps/sample", "20 (fixed)"),
             ("Bugs", "hashSimplex + acceptance always 1")]
else:
    info += [("Mode", "Paper (size 2-4)"),
             ("Features", "Corrected classifier, proper M-H, multi-chain")]
for label, val in info:
    sc(ws4.cell(row=row, column=1, value=label), BOLD)
    sc(ws4.cell(row=row, column=2, value=val), NORMAL, align=Alignment(horizontal="left", vertical="center"))
    row += 1

row += 1
hdrs = ["Type", "Exact", f"{METHOD_LABEL} Mean", "Std", "Mean Error"]
if has_ci: hdrs.append("CI Coverage")
for ci, h in enumerate(hdrs):
    sc(ws4.cell(row=row, column=1+ci, value=h), HEADER_FONT, HEADER_FILL)

for t in active_types:
    row += 1
    freqs = [r["freq"][t] for r in runs]
    mf = sum(freqs) / len(freqs)
    sf = (sum((f - mf)**2 for f in freqs) / len(freqs)) ** 0.5
    me = abs(mf - exact["freq"][t])
    sc(ws4.cell(row=row, column=1, value=f"Type {t+1}"), BOLD)
    sc(ws4.cell(row=row, column=2, value=exact["freq"][t]), NORMAL, EXACT_FILL, fmt="0.000000")
    sc(ws4.cell(row=row, column=3, value=mf), NORMAL, RUN_FILL, fmt="0.000000")
    sc(ws4.cell(row=row, column=4, value=sf), NORMAL, fmt="0.000000")
    sc(ws4.cell(row=row, column=5, value=me), NORMAL, ERR_FILL if me > 0.01 else None, fmt="0.000000")
    if has_ci:
        cc = sum(1 for r in runs if r["ci_lo"][t] <= exact["freq"][t] <= r["ci_hi"][t])
        cp = 100.0 * cc / NUM_RUNS
        sc(ws4.cell(row=row, column=6, value=f"{cp:.0f}%"), NORMAL,
           PatternFill("solid", fgColor="C6EFCE") if cp >= 90 else ERR_FILL)

row += 2
ame = [abs(sum(r["freq"][t] for r in runs)/len(runs) - exact["freq"][t]) for t in active_types]
sc(ws4.cell(row=row, column=1, value="Overall RMSE"), BOLD)
sc(ws4.cell(row=row, column=2, value=(sum(e**2 for e in ame)/len(ame))**0.5), NORMAL, fmt="0.000000")
row += 1
sc(ws4.cell(row=row, column=1, value="Overall MAE"), BOLD)
sc(ws4.cell(row=row, column=2, value=sum(ame)/len(ame)), NORMAL, fmt="0.000000")
if has_ci:
    row += 1
    oci = 100.0 * total_covered / total_checks if total_checks > 0 else 0
    sc(ws4.cell(row=row, column=1, value="Overall CI Coverage"), BOLD)
    sc(ws4.cell(row=row, column=2, value=f"{oci:.1f}%"), NORMAL,
       PatternFill("solid", fgColor="C6EFCE") if oci >= 90 else ERR_FILL)
row += 1
sc(ws4.cell(row=row, column=1, value=f"Avg Time (ms)"), BOLD)
sc(ws4.cell(row=row, column=2, value=sum(r["time_ms"] for r in runs)/NUM_RUNS), NORMAL, fmt="0.0")
aw(ws4)

# === SAVE ===
output_path = os.path.join(PROJECT_DIR, "results",
    f"sfd_exact_vs_{METHOD}_{DATASET}_{NUM_RUNS}runs.xlsx")
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"\nExcel saved: {output_path}")